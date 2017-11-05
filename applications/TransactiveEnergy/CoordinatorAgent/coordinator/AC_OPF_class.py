# -----------------------------------------------------------------------------
# Created by Xiangqi Zhu on Jun 16, 2016
# To run GAMS from python automatically and change the load data every time
# To read in all the data to a list 
# Change P, Q, and generator status
# Modified by xinda ke in results output section in 10/5/2016
# Code is modified by Quan Nguyen in July 2017
# -----------------------------------------------------------------------------


# -----------------------------------------------------------------------------
# For GAMS API
from gams import *
from gdxcc import *
import os
import openpyxl
from openpyxl.compat import range
import numpy
import pandas

from Export_Solution_class import Export_Solution 
# -----------------------------------------------------------------------------


class AC_OPF:     
    def __init__(self, Nbend, Nbend2, fileorder, simuyear, simutimestep, startrow, definesolarlocation, sum_bus_num, sum_load_num, sum_gennum, sum_distriGen,\
                    solar_curtail_busid, control_load_BusID,solar_qlimit_busID, load_bus_voltage_goal, solar_curtailment_off, demand_response_off, solar_q_off, shunt_switching_off,\
                    shuntSwitchingPenalty1, demand_response_decrease_penalty, demand_response_increase_penalty, solar_curtail_penalty, generator_voltage_deviation_penalty,\
                    load_bus_volt_pen_coeff_1, load_bus_volt_pen_coeff_2, load_bus_volt_dead_band, opt_Vgen_dev, opt_Vload_dev, previous_solution_as_start_point, icset, max_iter,\
                    bus_list_file, solar_index_file, solar_oneyear_file, GAMS_OptionSolution_GDXfile, GAMS_OPF_file, Feeder_Constraints_file,\
                    export_to_spreadsheet, system_folder,GDXcaseName, TimeStamp, feederLoad, P, Q, DRquantity, DRprice):
                 
        self.P                                       = P
        self.Q                                       = Q
        self.DRquantity                              = DRquantity
        self.DRprice                                 = DRprice
        
        self.feederLoad                              = feederLoad
        
        
        self.Nbend                                   = Nbend
        self.Nbend2                                  = Nbend2
        self.fileorder                               = fileorder
        self.simuyear                                = simuyear
        self.simutimestep                            = simutimestep
        
        # Define start row
        self.startrow                                = startrow 
        
        # Define the location of solar bus and solar apparent power
        self.definesolarlocation                     = definesolarlocation 
        
        self.pv_num                                  = sum_distriGen
        self.sum_bus_num                             = sum_bus_num  
        self.sum_load_num                            = sum_load_num 
        self.sum_gennum                              = sum_gennum  
        self.sum_distriGen                               = sum_distriGen  
        self.solar_curtail_busid                     = solar_curtail_busid  
        self.control_load_BusID                         = control_load_BusID  
        self.load_bus_voltage_goal                   = load_bus_voltage_goal             
        self.solar_curtailment_off                   = solar_curtailment_off    
        self.demand_response_off                     = demand_response_off    
        self.solar_q_off                             = solar_q_off         
        self.shunt_switching_off                     = shunt_switching_off     
        # Penalty terms
        self.shuntSwitchingPenalty1                  = shuntSwitchingPenalty1 
        self.demand_response_decrease_penalty        = demand_response_decrease_penalty 
        self.demand_response_increase_penalty        = demand_response_increase_penalty
        self.solar_curtail_penalty                   = solar_curtail_penalty  
        self.generator_voltage_deviation_penalty     = generator_voltage_deviation_penalty  
        self.load_bus_volt_pen_coeff_1               = load_bus_volt_pen_coeff_1  
        self.load_bus_volt_pen_coeff_2               = load_bus_volt_pen_coeff_2       
        self.load_bus_volt_dead_band                 = load_bus_volt_dead_band
        self.opt_Vgen_dev                            = opt_Vgen_dev
        self.opt_Vload_dev                           = opt_Vload_dev
        self.previous_solution_as_start_point        = previous_solution_as_start_point  
        self.icset                                   = icset
        self.max_iter                                = max_iter
        self.bus_list_file                           = bus_list_file
        self.solar_index_file                        = solar_index_file
        self.solar_oneyear_file                      = solar_oneyear_file
        self.GAMS_OptionSolution_GDXfile             = GAMS_OptionSolution_GDXfile
        self.GAMS_OPF_file                           = GAMS_OPF_file
        self.Feeder_Constraints_file                 = Feeder_Constraints_file
        
        self.export_to_spreadsheet                   = export_to_spreadsheet

        self.system_folder                           = system_folder        
        self.TimeStamp                               = TimeStamp
        self.GDXcaseName                             = GDXcaseName
        self.Display_Input()
        returnVal = self.Main_Program()
        
        self.returnVal = returnVal
        

    def Display_Input(self):
        # display input options
        print('num_time_steps: %d'                      % self.Nbend)
        print('shuntSwitchingPenalty: %f'               % self.shuntSwitchingPenalty1)
        print('load_bus_voltage_goal: %f'               % self.load_bus_voltage_goal)
        print('solar_curtailment_off: %d'               % self.solar_curtailment_off)
        print('demand_response_off: %d'                 % self.demand_response_off)
        print('solar_q_off: %d'                         % self.solar_q_off)
        print('shunt_switching_off: %d'                 % self.shunt_switching_off)
        print('load_bus_volt_pen_coeff_1: %f'           % self.load_bus_volt_pen_coeff_1)
        print('load_bus_volt_pen_coeff_2: %f'           % self.load_bus_volt_pen_coeff_2)
        print('load_bus_volt_dead_band: %f'             % self.load_bus_volt_dead_band)
        print('Control generator voltages: %f'          % self.opt_Vgen_dev) 
        print('Control load voltages: %f'               % self.opt_Vload_dev)         
        print('previous_solution_as_start_point: %u'    % self.previous_solution_as_start_point)
        print('file_number: %u'                         % self.fileorder)
        print('start_row_number: %u'                    % self.startrow)



    def Main_Program(self):
        #initialize the row list I'd like to read
        if self.startrow != 1:
            self.Nfinal  = self.startrow + self.Nbend
            self.rowlist = list(range(self.startrow, self.Nfinal))
            self.rowlist.insert(0,1)
        else:
            self.Nfinal  = self.Nbend+2
            self.rowlist = list(range(self.startrow, self.Nfinal))
            
    
        
        """ Read in input data into lists including the header """
        self.wb = openpyxl.load_workbook(self.system_folder + '/' + 'input_update' + str(self.fileorder) + '.xlsx')
        
#         self.p = [7, 18, 57, 1.405808599096, 1.3230449957, 1.681560294004, 1.405808599096, 1.3230449957, 1.681560294004, 1.405808599096, 1.3230449957, 1.681560294004]
        
        self.p  = self.P #self.Read_in(self.wb, 'P'   , self.sum_load_num, self.rowlist)
        self.q  = self.Q #self.Read_in(self.wb, 'Q'   , self.sum_load_num, self.rowlist)
        
        # comment out ------------------------------------------------------------------
#         self.p  = self.Read_in(self.wb, 'P'   , self.sum_load_num, self.rowlist)
#         self.q  = self.Read_in(self.wb, 'Q'   , self.sum_load_num, self.rowlist)
        # ------------------------------------------------------------------------------
        
        self.st = self.Read_in(self.wb, 'GenS', self.sum_gennum  , self.rowlist)
#         self.ip = self.Read_in(self.wb, 'GenP', self.sum_gennum  , self.rowlist)
        self.pv = self.Read_in(self.wb, 'pv'  , self.pv_num      , self.rowlist)   
        
        self.ip = [1, self.feederLoad, self.feederLoad, self.feederLoad]     # Define substation real power based on given pre-run data and the scale factor
#         self.ip = [1, 0.1, 0.1, 0.1]
           
        if self.definesolarlocation:
            self.solarindex  = self.Read_solar(self.system_folder + '/' + self.solar_index_file  , self.sum_distriGen, 3, 3)
        
            
        self.N               = self.Nbend * self.sum_gennum
        self.iq              = [0] * self.sum_gennum * len(self.rowlist)
        if self.demand_response_off==0:
            
            self.wb              = openpyxl.load_workbook(self.system_folder + '/' + self.Feeder_Constraints_file)
            self.sheet1          = self.wb.get_sheet_by_name('DR_Boundary')
            
            self.DemandRes_Price = self.DRprice
            self.DemandRes_load = self.DRquantity
            
            # comment out --------------------------------------------------------------
#             self.DemandRes_Price   = [None] * self.Nbend2
#             self.DemandRes_load   = [None] * self.Nbend2
#     
#             for icol in range(2, self.Nbend2+2):
#                 self.DemandRes_Price[icol-2]   = self.sheet1.cell(row=1, column=icol).value
#                 self.DemandRes_load[icol-2] = self.sheet1.cell(row=1, column=icol).value
#             for icol in range(2, self.Nbend2+2):
#                 self.DemandRes_Price.append(self.sheet1.cell(row=3, column=icol).value)
#                 self.DemandRes_load.append(self.sheet1.cell(row=2, column=icol).value)

            # --------------------------------------------------------------------------
                    
        [self.token1_vm, self.token2_Reac_P, self.token3_Reac_Solar, self.token4_tot_cost, self.token5_tot_loss, self.token6_tot_swit, self.token7_demand_res,\
        self.token8_solver_stat, self.token9_model_stat, self.token4_Distribu_Gen, self.token11_DR_inc_P, self.token12_DR_dec_P, self.token13_DR_inc_Q, self.token14_DR_dec_Q,\
        self.pgen, self.demand, self.generator_voltage, self.nongenerator_voltage, self.genvoltage_deviation, self.loadvoltage_deviation, self.loadvoltage_deviation_db,\
        self.Vschedule, self.genatbus, self.loadindex, self.solaratbus, self.rowLoadnongenatBus,\
        self.solarQlimit_Up, self.solarQlimit_Down, self.bus_va_degrees, self.switched_shunt_susceptance_values, self.switched_shunt_susceptance_by_bus_values]\
        = self.Lists_Init(31)
        
        self.bus_list_df             = pandas.read_excel(self.system_folder + '/' + self.bus_list_file)
        self.busindex                = list(self.bus_list_df.iloc[1:, 0])
        self.busindex                = list(map(int, self.busindex))
        self.genindex                = list(range(1, self.sum_gennum+1)) 

        
        self.Nb = 0
        
        # Print inputs of ACOPF
        # ------------------------- Test -------------------------------------------
#         self.p = [7, 18, 57, 1.33808817432, 1.69949922163, 2.91930600931, 1.33808817432, 1.69949922163, 2.91930600931, 1.33808817432, 1.69949922163, 2.91930600931]
#         self.q = [7, 18, 57, 0.781933400941, 0.503132155659, 1.04441352255, 0.781933400941, 0.503132155659, 1.04441352255, 0.781933400941, 0.503132155659, 1.04441352255]
#         self.DemandRes_Price = [1, 2, 3, 4, 5, 0, 4.90188634973, 9.50369900228, 14.0183879718, 18.2996843444]
#         self.DemandRes_load = [1, 2, 3, 4, 5, 0, 0.0462604737722, 0.0925209475445, 0.138781421317, 0.185041895089]
#         self.feederLoad = 5.30274177032
#         self.ip = [1, self.feederLoad, self.feederLoad, self.feederLoad]
        # ----------------------------------------------------------------------------

        print '[%s]' % ', '.join(map(str, self.p))
        print '[%s]' % ', '.join(map(str, self.q))
        print '[%s]' % ', '.join(map(str, self.DemandRes_Price))
        print '[%s]' % ', '.join(map(str, self.DemandRes_load))
        print '%s' % self.feederLoad
        print('end of printing inputs')
        
        """ MAIN ITERATION -------------------------------------------------------- """
        ACOPF_solved = False
        while self.Nb < self.Nbend:  
            self.Nb      += 1       
            
            """ CREATE DGX FILES -------------------------------------------------- """
            # Create gdx file of demand pd 
            self.CreateGDXfile(self.p, "UC_OPF_load_elwi_BV.gdx", "Run_newOPF_oneyear_simulation_2", "Pd", "RealPower", 1, GMS_DT_PAR, self.sum_load_num)
            # Create gdx file of Reactive demand q load
            self.CreateGDXfile(self.q, "UC_OPF_load_elwi_Q.gdx", "Run_newOPF_oneyear_simulation_2", "Qd", "ReactivePower", 1, GMS_DT_PAR, self.sum_load_num)    
            # Create gdx file of Generator Status
            self.CreateGDXfile(self.st, "geni_st.gdx", "Run_newOPF_oneyear_simulation_2", "status", "generator status", 1, GMS_DT_PAR, self.sum_gennum)
            # Create gdx file of scheduled real  power 
            self.CreateGDXfile(self.ip, "geni_p.gdx", "Run_newOPF_oneyear_simulation_2", "PgSch", "ScheduleRealPower", 1, GMS_DT_PAR, self.sum_gennum)    
            # Create gdx file of scheduled reactive  power
            self.CreateGDXfile(self.iq, "geni_q.gdx", "Run_newOPF_oneyear_simulation_2", "Qg", "ScheduleReactivePower", 1, GMS_DT_PAR, self.sum_gennum)    
            # Create gdx file of PV
            self.CreateGDXfile(self.pv, "Solar_real.gdx", "Run_newOPF_oneyear_simulation_2", "solar_r", "solar", 1, GMS_DT_PAR, self.pv_num)
            # Create gdx file for solar just ONCE
            if (self.definesolarlocation) & (self.Nb == 1):
                self.CreateGDX_solar(self.solarindex, "solarbus2.gdx", "Run_newOPF_oneyear_simulation_2", "solarbus", "solar", 1, GMS_DT_SET, self.sum_distriGen, False)
                self.CreateGDX_solar(self.solarindex, "solarlocation.gdx", "Run_newOPF_oneyear_simulation_2", "solarlocation", "solar", 2, GMS_DT_PAR, self.sum_distriGen, True)
            if (self.demand_response_off==0) & (self.Nb == 1):
                self.CreateGDXfile(self.DemandRes_load, "demandresponseload.gdx", "Run_newOPF_oneyear_simulation_2", "demLoad", "load", 1, GMS_DT_PAR, self.Nbend2)
                self.CreateGDXfile(self.DemandRes_Price, "demandresponseprice.gdx", "Run_newOPF_oneyear_simulation_2", "demPrice", "price", 1, GMS_DT_PAR, self.Nbend2)
            """ CALL GAMS TO SOLVE THE OPTIMIZATION PROBLEM ----------------------- """
            self.opt_time           = 0
            self.ws                 = GamsWorkspace()
            self.ws.__init__(
                working_directory   = os.getcwd(),
                system_directory    = None,
                #debug               = DebugLevel.ShowLog
                debug               = DebugLevel.Off
                )   
            t1 = self.ws.add_job_from_file(self.GAMS_OPF_file)
            # this is how to pass "--" options to the gams model. especially useful for --case and penalty parameters
            gams_options                                                = self.ws.add_options()
            gams_options.gdx                                            = self.GAMS_OptionSolution_GDXfile
            gams_options.defines["shuntSwitchingPenalty1"]              = str(self.shuntSwitchingPenalty1)
            gams_options.defines["demand_response_off"]                 = str(self.demand_response_off)
            gams_options.defines["solar_curtailment_off"]               = str(self.solar_curtailment_off)
            gams_options.defines["solar_q_off"]                         = str(self.solar_q_off)
            gams_options.defines["shunt_switching_off"]                 = str(self.shunt_switching_off)
            gams_options.defines["demand_response_decrease_penalty"]    = str(self.demand_response_decrease_penalty)
            gams_options.defines["demand_response_increase_penalty"]    = str(self.demand_response_increase_penalty)
            gams_options.defines["solar_curtail_penalty"]               = str(self.solar_curtail_penalty)
            gams_options.defines["generator_voltage_deviation_penalty"] = str(self.generator_voltage_deviation_penalty)
            gams_options.defines["load_bus_volt_pen_coeff_1"]           = str(self.load_bus_volt_pen_coeff_1)
            gams_options.defines["load_bus_volt_pen_coeff_2"]           = str(self.load_bus_volt_pen_coeff_2)
            gams_options.defines["load_bus_volt_dead_band"]             = str(self.load_bus_volt_dead_band)
            gams_options.defines["opt_Vgen_dev"]                        = str(self.opt_Vgen_dev)
            gams_options.defines["opt_Vload_dev"]                       = str(self.opt_Vload_dev)            
            gams_options.defines["previous_solution_as_start_point"]    = str(self.previous_solution_as_start_point)
            gams_options.defines["ic"]                                  = str(self.icset[self.opt_time])
            gams_options.defines["case"]                                = str(self.GDXcaseName)
            gams_options.defines["control_solarQ"]                      = str("1")
            # Run GAMS
            # NOTE:  DO NOT put t1 as an attribute of object SELF; otherwise, you MUST restart Python if you want to run the code again. The command "os.remove" DOES NOT help
            #return # uncomment this line to quit right before the first gams solve
            try:
                t1.run(gams_options)  
                if (t1.out_db["mstat"].find_record().value == 2):
                    ACOPF_solved = True 
                    print('ACOPF is solved') 
                else:
                    ACOPF_solved = False
                    print('Could not solve ACOPF - mstat is 5.0') 
            except:
                print('Could not solve ACOPF')
                 
            # Check the status of the model and solution
            print ('mstat: %s' %t1.out_db["mstat"].find_record().value)
            print ('Opt time: %s' %self.opt_time)
            
            while (t1.out_db["mstat"].find_record().value != 2 and self.opt_time < self.max_iter):
                self.opt_time += 1
                
                gams_options.defines["ic"] = str(self.icset[self.opt_time])
                # Run GAMS again
                try:
                    if (t1.out_db["mstat"].find_record().value == 2):
                        ACOPF_solved = True 
                        print('ACOPF is solved') 
                    else:
                        ACOPF_solved = False
                        print('Could not solve ACOPF - mstat is 5.0')
                except:
                    print('Could not solve ACOPF')
            
                print ('mstat: %s' %t1.out_db["mstat"].find_record().value)
                print ('Opt time: %s' %self.opt_time)
                
            print ('Step: %s' %self.Nb)
            
   
            """ COLLECT THE RESULT FROM GAMS --------------------------------------- """
            if (ACOPF_solved == True):
                # Collect the headers      
                self.switched_shunt_susceptance_keys     = [(rec.key(0), rec.key(1)) for rec in t1.out_db["switched_shunt_susceptance_keys"]] 
                self.switched_shunt_susceptance_buses    = [rec[0] for rec in self.switched_shunt_susceptance_keys]  
           
                if (self.Nb == 1): # only to collect the header information ONCE
                    for comp in t1.out_db["atBus"]:
                        self.genatbus.append(int(comp.key(1)))    
                    for comp in t1.out_db["Pd"]:
                        self.loadindex.append(int(comp.key(0)))     
                    for comp in t1.out_db["VoltageDevNongen"]:                    
                        self.rowLoadnongenatBus.append(int(comp.key(0)))
                    for comp in t1.out_db["type"]:                    
                        if comp.value == 3:
                            self.slackbus = int(comp.key(0))                  
                    # Number of nongen buses connected to loads
                    self.len_nongenLoad = len(self.rowLoadnongenatBus)                    
                    # Non-gen voltage deadband (parameter)
                    self.loadvoltage_deviation_db        = self.Get_Result(t1, self.loadvoltage_deviation_db, "load_bus_volt_dead_band", 2, 1)                 
                
                # Voltage magnitude at generator bus (parameters)
                self.token1_vm                           = self.Get_Result(t1, self.token1_vm, "Vm", 2, self.sum_bus_num)
                # Reactive power generated by generators (variables)
                self.token2_Reac_P                       = self.Get_Result(t1, self.token2_Reac_P, "V_Q", 1, self.sum_gennum)
                # Reactive power generated by solars and the associated inverters (variables)
                self.token3_Reac_Solar                   = self.Get_Result(t1, self.token3_Reac_Solar, "Q_S", 1, self.sum_distriGen)
                
                
                
                
                 
                # Reactive power generated by solars and the associated inverters (variables)
                #distributed generation
                self.token4_Distribu_Gen                = self.Get_Result(t1, self.token4_Distribu_Gen, "P_S", 1, self.sum_distriGen)    
                # Real power generated by generators (Variables)
                self.pgen                                = self.Get_Result(t1, self.pgen, "V_P", 1, self.sum_gennum)       
                # Demand (Parameters)
                self.demand                              = self.Get_Result(t1, self.demand, "Pd", 2, self.sum_load_num)
                # Generator voltages (Parameters)
                self.generator_voltage                   = self.Get_Result(t1, self.generator_voltage, "vmgenBus", 2, self.sum_gennum) 
                # Generator voltages schedule (Parameters)
                self.Vschedule                           = self.Get_Result(t1, self.Vschedule, "Vsch", 2, self.sum_gennum)
                # Load voltages (Parameters)
                self.nongenerator_voltage                = self.Get_Result(t1, self.nongenerator_voltage, "vmnongen", 2, self.sum_bus_num-self.sum_gennum) 
                # Non-gen load voltage deviation (parameters)
                self.loadvoltage_deviation               = self.Get_Result(t1, self.loadvoltage_deviation, "VoltageDevNongen", 2, self.len_nongenLoad)            
                # Values of shunt susceptances (Parameters)
                self.switched_shunt_susceptance_values   = self.Get_Result_shunt(t1, self.switched_shunt_susceptance_values, "switched_shunt_susceptance_final")  
                # Values of shunt susceptances (Parameters)
                self.bus_va_degrees                      = self.Get_Result(t1, self.bus_va_degrees, "Va", 2, self.sum_bus_num)  
                self.bus_va_degrees.insert((self.Nb-1)*self.sum_bus_num + self.slackbus-1 , 0) # insert zero angle of the slack bus
                # Total cost
                self.token4_tot_cost                     = self.Get_Result(t1, self.token4_tot_cost, "total_cost", 3, 1)
                # Total lost
                self.token5_tot_loss                     = self.Get_Result(t1, self.token5_tot_loss, "TotalLoss", 3, 1)    
                # Total switch cost
                self.token6_tot_swit                     = self.Get_Result(t1, self.token6_tot_swit, "TotalSwitches", 3, 1)     
              
                
                
                
                # Load voltage deviation
                #controllable load shift
#                 self.token7_demand_res               = self.Get_Result(t1, self.token7_demand_res, "V_dem_Load", 1, 5)
                self.token7_demand_res               = list(numpy.array(self.Get_Result(t1, self.token7_demand_res, "V_dem_Load", 1, 5))*t1.out_db["baseMVA"].find_record().value) 
                # sstat
                self.token8_solver_stat                  = self.Get_Result(t1, self.token8_solver_stat, "sstat", 3, 1)    
                # mstat
                self.token9_model_stat                   = self.Get_Result(t1, self.token9_model_stat, "mstat", 3, 1)      
            
        
        if (ACOPF_solved == True):
            ## EXTRA calculations
            # Calculate voltage deviation at generators buses                    
            self.genvoltage_deviation                    = list(numpy.array(self.generator_voltage) - numpy.array(self.Vschedule))
            #self.genvoltage_deviation_per                = list(numpy.array(self.genvoltage_deviation) / numpy.array(self.Vschedule) * 100)
            self.genvoltage_deviation_per=self.genvoltage_deviation
            # Calculate voltage deviation at nongen load buses
            self.loadvoltage_deviation_per             = [0] * self.len_nongenLoad * self.Nbend
            for idx in range(len(self.loadvoltage_deviation)):
                if abs(self.loadvoltage_deviation[idx]) < self.loadvoltage_deviation_db[0]:
                    self.loadvoltage_deviation_per[idx] = 0
                elif self.loadvoltage_deviation[idx] > self.loadvoltage_deviation_db[0]:
                    self.loadvoltage_deviation_per[idx] = (self.loadvoltage_deviation[idx] - self.loadvoltage_deviation_db[0]) / 1 * 100
                elif self.loadvoltage_deviation[idx] < (-self.loadvoltage_deviation_db[0]):
                    self.loadvoltage_deviation_per[idx] = abs(self.loadvoltage_deviation[idx] + self.loadvoltage_deviation_db[0]) / 1 * 100
            
            """ Return values to the main function """
            returnVal = dict()
            returnVal['solved'] = True
            returnVal['DER'] = [x * t1.out_db["baseMVA"].find_record().value for x in self.token4_Distribu_Gen]
            returnVal['DRquantity'] = sum(self.token7_demand_res)
            returnVal['substationP'] = self.pgen[0] * t1.out_db["baseMVA"].find_record().value
            returnVal['SocialWelfare'] = self.token4_tot_cost[0]
        
        else:
            returnVal = dict()
            returnVal['solved'] = False
        
        return returnVal

    def Read_in(self, wb, sheet_name, object_num, row_list):
        sheet = wb.get_sheet_by_name(sheet_name)
        ans   = [None] * object_num * len(row_list)   
        i     = 0
        if self.TimeStamp:
            start_col = 2
        else:
            start_col = 1            
        for row_numb in row_list:
            for colum_num in range(start_col, object_num+start_col):
                ans[i] = sheet.cell(row=row_numb, column=colum_num).value
                i     += 1
        return ans        
    

    def Read_solar(self, in_file, sum_distriGen, row, kmult):
        wb    = openpyxl.load_workbook(in_file)
        sheet = wb.active
        #TO be checked
        ans   = [None] * kmult * sum_distriGen
        i     = 0
        for row_numb in range(1, row+1):
            for colum_num in range(2, sum_distriGen+2):
                ans[i] = sheet.cell(row=row_numb, column=colum_num).value
                i     += 1
        return ans            
    
    
    def Lists_Init(self, n):
        ans = [[] for x in range(n)]
        return ans    
    
    
    def CreateGDXfile(self, input_list, file1, file2, kw1, kw2, str_dim, gms_kw, set_num):
        gdxHandle = new_gdxHandle_tp()
        rc        = gdxCreate(gdxHandle, GMS_SSSIZE)
        assert rc[0], rc[1]
        assert gdxOpenWrite(gdxHandle, file1, file2)[0]
        assert gdxDataWriteStrStart(gdxHandle, kw1, kw2, str_dim, gms_kw , 0)
        
        values  = doubleArray(GMS_VAL_MAX)
        t       = set_num * self.Nb
        tt      = 0
        while t <= (set_num*self.Nb + set_num-1):
            values[GMS_VAL_LEVEL] = input_list[t]
            gdxDataWriteStr(gdxHandle, [str(input_list[tt])], values)
            t  += 1
            tt += 1
    
        assert gdxDataWriteDone(gdxHandle)
        assert not gdxClose(gdxHandle)
        assert gdxFree(gdxHandle)
    
        
    def CreateGDX_solar(self, input_list, file1, file2, kw1, kw2, str_dim, gms_kw, set_num, Boo):
        gdxHandle = new_gdxHandle_tp()
        rc        = gdxCreate(gdxHandle, GMS_SSSIZE)
        assert rc[0], rc[1]
        assert gdxOpenWrite(gdxHandle, file1, file2)[0]
        assert gdxDataWriteStrStart(gdxHandle, kw1, kw2, str_dim, gms_kw , 0)
        
        values  = doubleArray(GMS_VAL_MAX)
        t       = 0
        while t <= (set_num-1):
            if Boo:
                values[GMS_VAL_LEVEL] = input_list[t + str_dim*set_num]
            else:
                values[GMS_VAL_LEVEL] = 1
            if str_dim == 1:          
                gdxDataWriteStr(gdxHandle, [str(input_list[t])], values)
            else:
                gdxDataWriteStr(gdxHandle, [str(input_list[t]), str(input_list[t+set_num])], values)
            t += 1
            
        assert gdxDataWriteDone(gdxHandle)
        assert not gdxClose(gdxHandle)
        assert gdxFree(gdxHandle)
        
    
    def Get_Result(self, ws, token, a_field, Type, length):
        if self.opt_time != self.max_iter:
            if Type == 1: # Variable
                for comp in ws.out_db[a_field]:
                    token.append(comp.level)
            if Type == 2: # Parameter   
                for comp in ws.out_db[a_field]:
                    token.append(comp.value)
            if Type == 3: # 1_dimentional Value
                token.append(ws.out_db[a_field].find_record().value)
        else:
            for idx in range(length):
                token.append(0)  
        return token
    
    
    def Get_Result_shunt(self, ws, token, a_field):
        if self.opt_time != self.max_iter:        
            keys        = [(rec.key(0), rec.key(1)) for rec in ws.out_db[a_field]]
            shuntbuses  = [rec[0] for rec in keys]       
            new_token = [0,] * len(self.switched_shunt_susceptance_buses)
            realshunt = []
            for comp in ws.out_db[a_field]:
                realshunt.append(comp.value)  
            count = 0              
            for idx1 in range(len(shuntbuses)):
                for idx2 in range(len(self.switched_shunt_susceptance_buses)):
                    if shuntbuses[idx1] == self.switched_shunt_susceptance_buses[idx2]:
                        new_token[idx2] = realshunt[count]
                        count+=1
                        break
            token += new_token
        else:
            for idx in range(len(self.switched_shunt_susceptance_buses)):
                token.append(0)
        return token    




    
