
# write by xinda ke in results output section in 10/5/2016

## For GAMS API
from __future__ import print_function
from gams import *
from gdxcc import *
import os
import sys
import time
import openpyxl
import subprocess
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import getopt
import output_class

#define some class over here





def output_solutionfile( firstrow, arg2 ):
    #    dest_filename = 'finalsolution'+ str(fileorder) + '.xlsx'
        #filenum=filenum+1
        #switched_shunt_unique_buses.sort() # these are strings anyway so just forget it


        from openpyxl import Workbook
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="V_Magnitude")
        ws2 = wb.create_sheet(title="Qg") 
        ws3 = wb.create_sheet(title="Qsolar") 
        ws4 = wb.create_sheet(title="totalcost")
        ws5 = wb.create_sheet(title="totalloss")
        ws6 = wb.create_sheet(title="totalswitches")
        ws7 = wb.create_sheet(title="averageloadbusvoltagedeviation")
        ws8 = wb.create_sheet(title="sstat")
        ws9 = wb.create_sheet(title="mstat")
        ws15 = wb.create_sheet(title="Pgen")
        ws16 = wb.create_sheet(title="demand")
        ws17 = wb.create_sheet(title="generator_voltage")
        ws18 = wb.create_sheet(title="non_generator_voltage")
        ws19 = wb.create_sheet(title="generator_V_deviation")
        ws20 = wb.create_sheet(title="nongenerator_V_deviation")
        ws24 = wb.create_sheet(title="switched_shunt_susceptance")

   #     ws25 = wb.create_sheet(title="switched_B_by_bus")
        ws26 = wb.create_sheet(title="bus_va_degrees")

        ws.append(list(firstrow.busindex))
        ws2.append(firstrow.genindex)
        ws2.append(firstrow.genatbus)
        ws3.append(firstrow.solaratbus)
        ws16.append(firstrow.loadindex)
        ws17.append(list(firstrow.genindex))
        ws15.append(list(firstrow.genindex))
        ws18.append(list(firstrow.busindex))
        ws17.append(firstrow.genatbus)
        ws15.append(firstrow.genatbus)
        ws18.append(firstrow.genatbus)
        ws24.append(firstrow.switched_shunt_susceptance_buses)
        ws24.append(firstrow.switched_shunt_susceptance_shunts)
    #    ws25.append(switched_shunt_unique_buses)
        ws26.append(list(firstrow.busindex))
        
        
  #     row+1)*sum_bus_num+1)])
            
        del ws
        del ws2
        del ws3
        del ws4
        del ws5
        del ws6
        del ws7
        del ws8
        del ws9
        del ws17
        del ws18
        del ws19
        del ws20
        del ws24
      #  del ws25
        del ws26
      #  wb.save(dest_filename)
        token1_vm=[]
        token2_Reac_P=[]
        token3_Reac_Solar=[]
        token4_tot_cost=[]
        token5_tot_loss=[]
        token6_tot_swit=[]
        token7_v_devi_loadbus=[]
        token8_solver_stat=[]
        token9_model_stat=[]
        token10_solar_p_curt=[]
        token11_DR_inc_P=[]
        token12_DR_dec_P=[]
        token13_DR_inc_Q=[]
        token14_DR_dec_Q=[]
        pgen=[]
        demand=[]
        generator_voltage=[]
        nongenerator_voltage=[]
        genvoltage_deviation=[]
        nongenvoltage_deviation=[]
        Vschedule=[]
        solarQlimit_Up=[]
        solarQlimit_Down=[]
        bus_va_degrees = []
        switched_shunt_susceptance_values=[]
        switched_shunt_susceptance_by_bus_values=[]
        
        
  


