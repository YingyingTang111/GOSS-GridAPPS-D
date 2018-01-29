# -----------------------------------------------------------------------------
# Created by Xiangqi Zhu on Jun 16, 2016
# To run GAMS from python automatically and change the load data every time
# To read in all the data to a list 
# Change P, Q, and generator status
# Modified by xinda ke in results output section in 10/5/2016
# Code is modified by Quan Nguyen in July 2017
# -----------------------------------------------------------------------------


# -----------------------------------------------------------------------------
import datetime
from openpyxl import Workbook
import pandas
# import matplotlib.pyplot as plt
# import matplotlib.dates as mdates
# -----------------------------------------------------------------------------



class Export_Solution:
    def __init__(self, Nb, Nbend, fileorder, busindex, genindex, genatbus, solaratbus, solarindex, loadindex, sum_load_num, sum_gennum, sum_solar, sum_bus_num,\
                 switched_shunt_susceptance_buses, simutimestep, simuyear, startrow, token1_vm, token2_Reac_P, token3_Reac_Solar, token4_tot_cost, token5_tot_loss,\
                 token6_tot_swit, token7_v_devi_loadbus, token8_solver_stat, token9_model_stat, pgen, demand,\
                 generator_voltage, nongenerator_voltage, genvoltage_deviation, loadvoltage_deviation,\
                 genvoltage_deviation_per, loadvoltage_deviation_per, rowLoadnongenatBus, len_nongenLoad,\
                 switched_shunt_susceptance_values, bus_va_degrees,\
                 export_to_spreadsheet): 
        self.Nb                                 = Nb
        self.Nbend                              = Nbend
        self.fileorder                          = fileorder
        self.busindex                           = busindex
        self.genindex                           = genindex
        self.genatbus                           = genatbus
        self.solaratbus                         = solaratbus
        self.loadindex                          = loadindex
        self.sum_load_num                       = sum_load_num 
        self.sum_gennum                         = sum_gennum
        self.sum_solar                          = sum_solar
        self.sum_bus_num                        = sum_bus_num
        self.rowshuntatBus                      = switched_shunt_susceptance_buses
        self.simutimestep                       = simutimestep
        self.simuyear                           = simuyear
        self.startrow                           = startrow
        
        self.token1_vm                          = token1_vm
        self.token2_Reac_P                      = token2_Reac_P
        self.token3_Reac_Solar                  = token3_Reac_Solar
        self.token4_tot_cost                    = token4_tot_cost
        self.token5_tot_loss                    = token5_tot_loss
        self.token6_tot_swit                    = token6_tot_swit
        self.token7_v_devi_loadbus              = token7_v_devi_loadbus
        self.token8_solver_stat                 = token8_solver_stat
        self.token9_model_stat                  = token9_model_stat
        self.pgen                               = pgen
        self.demand                             = demand
        self.generator_voltage                  = generator_voltage
        self.nongenerator_voltage               = nongenerator_voltage
        self.genvoltage_deviation               = genvoltage_deviation
        self.genvoltage_deviation_per           = genvoltage_deviation_per
        self.loadvoltage_deviation              = loadvoltage_deviation
        self.loadvoltage_deviation_per          = loadvoltage_deviation_per
        self.rowLoadnongenatBus                 = rowLoadnongenatBus
        self.len_nongenLoad                     = len_nongenLoad
        self.switched_shunt_susceptance_values  = switched_shunt_susceptance_values
        self.bus_va_degrees                     = bus_va_degrees
        self.solarindex                         = solarindex
        self.export_to_spreadsheet              = export_to_spreadsheet
        
        if self.export_to_spreadsheet:
            self.Solution_to_Spreadsheet()

                        

        
    def Solution_to_Spreadsheet(self):
        """ WRITE THE DATA INTO A SPREADSHEET ----------------------------- """
        ## Create title for each sheet in the spreadsheet containing the result
        if (self.Nb == self.Nbend):
            dest_filename           = 'finalsolution'+ str(self.fileorder) + '.xlsx'
            wb                      = Workbook(write_only   = True)
            ws_Vm                   = wb.create_sheet(title = "V_Magnitude")
            ws_Qg                   = wb.create_sheet(title = "Qg") 
            ws_Qsolar               = wb.create_sheet(title = "Qsolar") 
            ws_Tcost                = wb.create_sheet(title = "totalcost")
            ws_Tloss                = wb.create_sheet(title = "totalloss")
            ws_Tswitches            = wb.create_sheet(title = "totalswitches")
            ws_Vdeviation           = wb.create_sheet(title = "averageloadbusvoltagedeviation")
            ws_sstat                = wb.create_sheet(title = "sstat")
            ws_mstat                = wb.create_sheet(title = "mstat")
            ws_Pgen                 = wb.create_sheet(title = "Pgen")
            ws_demand               = wb.create_sheet(title = "demand")
            ws_GenVm                = wb.create_sheet(title = "generator_voltage")
            ws_NonGenVm             = wb.create_sheet(title = "non_generator_voltage")
            ws_GenVdeviation        = wb.create_sheet(title = "generator_V_deviation")
            ws_GenVdeviation_per    = wb.create_sheet(title = "generator_V_deviation_per")            
            ws_Loaddeviation        = wb.create_sheet(title = "Load_V_deviation")
            ws_Loaddeviation_per    = wb.create_sheet(title = "Load_V_deviation_per")            
            ws_swShunt              = wb.create_sheet(title = "switched_shunt_susceptance")
            ws_BusVAngle            = wb.create_sheet(title = "bus_va_degrees")
            
            # Add the header for each row in each sheet
            rowbusid                = list(self.busindex)
            rowbusid.insert(0, 'Bus ID')
            rowgenid                = list(self.genindex)
            rowgenid.insert(0,'Gen ID')
            rowgenatBus             = self.genatbus
            rowgenatBus.insert(0,'Gen at bus')
            rowsolaratBus           = self.solaratbus
            rowsolaratBus.insert(0,'Solar at bus')
            rowsolarindex           = self.solarindex
            rowsolarindex.insert(0,'Solar index')
            rowloadbusindex         = self.loadindex
            rowloadbusindex.insert(0,'Load at bus index')
            rowloadid               = list(range(1,self.sum_load_num+1))
            rowloadid.insert(0,'Load index') 
            rowLoadnongenatBus      = self.rowLoadnongenatBus
            rowLoadnongenatBus.insert(0,'Load-NonGen at bus')            
            rowloadnongenid         = list(range(1, self.len_nongenLoad+1))
            rowloadnongenid.insert(0,'Load-NonGen ID') 
            rownongenatBus          = list(set(rowbusid[1:]) - set(rowgenatBus[1:]))            
            rownongenatBus.insert(0,'NonGen at bus')   
            rownongenid             = list(range(1, len(rownongenatBus)))            
            rownongenid.insert(0,'NonGen ID')             
            rowshuntid              = list(range(1, len(self.rowshuntatBus)+1))
            rowshuntid.insert(0, 'Shunt index')
            rowshuntatBus           = self.rowshuntatBus
            rowshuntatBus.insert(0, 'Shunt at bus')              
            emptyrow                    = ['Time step']
            ws_Vm                       = self.Add_Title(ws_Vm,                     [rowbusid, emptyrow, emptyrow])
            ws_Qg                       = self.Add_Title(ws_Qg,                     [rowgenatBus, rowgenid, emptyrow])
            ws_Qsolar                   = self.Add_Title(ws_Qsolar,                 [rowsolaratBus, rowsolarindex, emptyrow])        
            ws_Tcost                    = self.Add_Title(ws_Tcost,                  [emptyrow] * 3)  
            ws_Tloss                    = self.Add_Title(ws_Tloss,                  [emptyrow] * 3)  
            ws_Tswitches                = self.Add_Title(ws_Tswitches,              [emptyrow] * 3)  
            ws_Vdeviation               = self.Add_Title(ws_Vdeviation,             [emptyrow] * 3)  
            ws_sstat                    = self.Add_Title(ws_sstat,                  [emptyrow] * 3) 
            ws_mstat                    = self.Add_Title(ws_mstat,                  [emptyrow] * 3) 
            ws_demand                   = self.Add_Title(ws_demand,                 [rowloadbusindex, rowloadid, emptyrow]) 
            ws_GenVm                    = self.Add_Title(ws_GenVm,                  [rowgenatBus, rowgenid, emptyrow]) 
            ws_GenVdeviation            = self.Add_Title(ws_GenVdeviation,          [rowgenatBus, rowgenid, emptyrow])
            ws_GenVdeviation_per        = self.Add_Title(ws_GenVdeviation_per,      [rowgenatBus, rowgenid, emptyrow])           
            ws_Pgen                     = self.Add_Title(ws_Pgen,                   [rowgenatBus, rowgenid, emptyrow])         
            ws_NonGenVm                 = self.Add_Title(ws_NonGenVm,               [rownongenatBus, rownongenid, emptyrow])     
            ws_Loaddeviation            = self.Add_Title(ws_Loaddeviation,          [rowLoadnongenatBus, rowloadnongenid, emptyrow]) 
            ws_Loaddeviation_per        = self.Add_Title(ws_Loaddeviation_per,      [rowLoadnongenatBus, rowloadnongenid, emptyrow])             
            ws_swShunt                  = self.Add_Title(ws_swShunt,                [rowshuntatBus, rowshuntid, emptyrow])
            ws_BusVAngle                = self.Add_Title(ws_BusVAngle,              [rowbusid, emptyrow, emptyrow])        
                   
            ## Write the result into a spreadsheet, including both time and result data
            current_time = datetime.datetime(self.simuyear, 1, 1, 0, 0)
            current_time = current_time+datetime.timedelta(days = self.fileorder-1) + datetime.timedelta(minutes = (self.startrow-1)*self.simutimestep)
            for irow in range(self.Nb):
                a_row = []
                a_row.append(str(current_time.strftime('%Y-%m-%d %H:%M:%S')))              
                ws_Vm                   = self.WriteResult(ws_Vm,                   a_row, self.token1_vm, irow, self.sum_bus_num)
                ws_Qg                   = self.WriteResult(ws_Qg,                   a_row, self.token2_Reac_P, irow, self.sum_gennum)
                ws_Qsolar               = self.WriteResult(ws_Qsolar,               a_row, self.token3_Reac_Solar, irow, self.sum_solar)
                ws_Tcost                = self.WriteResult(ws_Tcost,                a_row, self.token4_tot_cost, irow, 1)
                ws_Tloss                = self.WriteResult(ws_Tloss,                a_row, self.token5_tot_loss, irow, 1)
                ws_Tswitches            = self.WriteResult(ws_Tswitches,            a_row, self.token6_tot_swit, irow, 1)
                ws_Vdeviation           = self.WriteResult(ws_Vdeviation,           a_row, self.token7_v_devi_loadbus, irow, 1)
                ws_sstat                = self.WriteResult(ws_sstat,                a_row, self.token8_solver_stat, irow, 1)
                ws_mstat                = self.WriteResult(ws_mstat,                a_row, self.token9_model_stat, irow, 1)
                ws_Pgen                 = self.WriteResult(ws_Pgen,                 a_row, self.pgen, irow, self.sum_gennum)
                ws_demand               = self.WriteResult(ws_demand,               a_row, self.demand, irow, self.sum_load_num)
                ws_GenVm                = self.WriteResult(ws_GenVm,                a_row, self.generator_voltage, irow, self.sum_gennum)
                ws_NonGenVm             = self.WriteResult(ws_NonGenVm,             a_row, self.nongenerator_voltage, irow, self.sum_bus_num-self.sum_gennum)
                ws_GenVdeviation        = self.WriteResult(ws_GenVdeviation,        a_row, self.genvoltage_deviation, irow, self.sum_gennum)
                ws_GenVdeviation_per    = self.WriteResult(ws_GenVdeviation_per,    a_row, self.genvoltage_deviation_per, irow, self.sum_gennum)                
                ws_Loaddeviation        = self.WriteResult(ws_Loaddeviation,        a_row, self.loadvoltage_deviation, irow, self.len_nongenLoad)
                ws_Loaddeviation_per    = self.WriteResult(ws_Loaddeviation_per,    a_row, self.loadvoltage_deviation, irow, self.len_nongenLoad)
                ws_swShunt              = self.WriteResult(ws_swShunt,              a_row, self.switched_shunt_susceptance_values, irow, len(self.rowshuntatBus)-1)
                ws_BusVAngle            = self.WriteResult(ws_BusVAngle,            a_row, self.bus_va_degrees, irow, self.sum_bus_num)
                current_time            = current_time + datetime.timedelta(minutes = self.simutimestep)
                
            ## Free the memory
            del ws_Vm, ws_Qg, ws_Qsolar, ws_Tcost, ws_Tloss, ws_Tswitches, ws_Vdeviation, ws_sstat, ws_mstat, ws_GenVm, ws_NonGenVm, ws_GenVdeviation, ws_Loaddeviation, ws_swShunt, ws_BusVAngle
            wb.save(dest_filename)
            
  
               

    
    
    def WriteResult(self, w_space, temp_row, token, idx_row, num_set):
        temp_row.extend(list(map(str, token[idx_row*num_set : ((idx_row+1)*num_set)])))            
        w_space.append(temp_row)
        del temp_row[1:]
        return w_space
    
    
    
    def Add_Title(self, a_ws, a_list_of_title):
        for a_title in a_list_of_title:
            a_ws.append(a_title)
        return a_ws  
    
